import pandas as pd
import io
import os
import openpyxl
from datetime import datetime

def export_orders_to_excel(orders_data_list):
    """
    Xuất danh sách đơn hàng ra file Excel bằng cách load template Nobita 
    để giữ nguyên format (styles, columns).
    """
    template_path = 'c:/app_theu/order-import-template.xlsx'
    
    # Load template
    if os.path.exists(template_path):
        wb = openpyxl.load_workbook(template_path)
    else:
        # Fallback nếu không thấy file template (tạo mới với headers chuẩn)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append([
            "Stt", "Khách hàng", "SĐT", "Người nhận", "Địa chỉ", 
            "Sản phẩm", "Mã sản phẩm", "Số lượng", "Trọng lượng (g)", "Giá", 
            "Giảm giá", "Loại Giảm Giá", "Nội dung để in", "Ghi chú nội bộ", 
            "Trả trước", "Chuyển khoản", "Tiền khách đưa", "Quẹt thẻ", 
            "Phí Vận Chuyển", "Hình thức thanh toán phí vận chuyển", "Nguồn đơn hàng"
        ])

    ws = wb.active
    
    # Xóa dữ liệu cũ từ hàng 2 trở đi (giữ lại header ở hàng 1)
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row)

    # Điền dữ liệu mới
    for idx, data in enumerate(orders_data_list):
        order = data['order_info']
        items = data['items']
        
        # 1. Ghép chuỗi tên sản phẩm (sp1 + sp2 + ...)
        product_names = [item.get('ten_sp', '').strip() for item in items if item.get('ten_sp')]
        full_product_str = " + ".join(product_names)
        
        ma_don = order.get('ma_don', '')
        ten_khach = order.get('ten_khach', '')
        
        # Map dữ liệu vào các cột (1-indexed)
        # 1: Stt
        ws.cell(row=idx+2, column=1, value=idx+1)
        # 2: Khách hàng (Mã đơn + tên)
        ws.cell(row=idx+2, column=2, value=f"{ma_don} {ten_khach}".strip())
        # 3: SĐT
        ws.cell(row=idx+2, column=3, value=order.get('sdt', ''))
        # 4: Người nhận (giống khách hàng)
        ws.cell(row=idx+2, column=4, value=f"{ma_don} {ten_khach}".strip())
        # 5: Địa chỉ
        ws.cell(row=idx+2, column=5, value=order.get('dia_chi', ''))
        # 6: Sản phẩm
        ws.cell(row=idx+2, column=6, value=full_product_str)
        # 7: Mã sản phẩm (để trống)
        # 8: Số lượng (cố định 1)
        ws.cell(row=idx+2, column=8, value=1)
        # 9: Trọng lượng (g) (cố định 500)
        ws.cell(row=idx+2, column=9, value=500)
        # 10: Giá (Thành tiền)
        ws.cell(row=idx+2, column=10, value=order.get('thanh_tien', 0))
        # 11: Giảm giá (trống)
        # 12: Loại Giảm Giá (trống)
        # 13: Nội dung để in (Ghi chú)
        ws.cell(row=idx+2, column=13, value=order.get('ghi_chu', ''))
        # 14: Ghi chú nội bộ (trống)
        # 15: Trả trước (trống)
        # 16: Chuyển khoản (Đã cọc)
        ws.cell(row=idx+2, column=16, value=order.get('da_coc', 0))
        # 17-19: Trống
        # 20: Hình thức thanh toán phí vận chuyển (Người gửi)
        ws.cell(row=idx+2, column=20, value="Người gửi")
        # 21: Nguồn đơn hàng (trống)

    # Xuất ra BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

